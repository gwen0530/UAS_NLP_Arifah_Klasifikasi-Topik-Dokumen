import sqlite3
import pandas as pd
import os
import json
from datetime import datetime
from config import DATABASE_PATH, DEFAULT_DATASET_PATH

def get_db():
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS Dataset (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            text TEXT NOT NULL,
            label TEXT NOT NULL,
            length INTEGER,
            word_count INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS PredictionHistory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            text TEXT NOT NULL,
            predicted_label TEXT NOT NULL,
            confidence REAL NOT NULL,
            probability_json TEXT NOT NULL,
            execution_time REAL NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS TrainingHistory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            accuracy REAL NOT NULL,
            precision REAL NOT NULL,
            recall REAL NOT NULL,
            f1_score REAL NOT NULL,
            dataset_size INTEGER NOT NULL,
            trained_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS UserActivity (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            action TEXT NOT NULL,
            details TEXT,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    
    cursor.execute('SELECT COUNT(*) as count FROM Dataset')
    row = cursor.fetchone()
    if row['count'] == 0 and os.path.exists(DEFAULT_DATASET_PATH):
        df = pd.read_csv(DEFAULT_DATASET_PATH)
        for _, r in df.iterrows():
            txt = str(r['text']).strip()
            lbl = str(r['label']).strip()
            cursor.execute(
                'INSERT INTO Dataset (text, label, length, word_count) VALUES (?, ?, ?, ?)',
                (txt, lbl, len(txt), len(txt.split()))
            )
        conn.commit()
        log_activity("Database Seed", f"Imported {len(df)} initial records from dataset.csv")
        
    conn.close()

def log_activity(action, details=""):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO UserActivity (action, details) VALUES (?, ?)',
            (action, details)
        )
        conn.commit()
        conn.close()
    except Exception:
        pass

def export_predictions_to_excel(output_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, text, predicted_label, confidence, execution_time, created_at FROM PredictionHistory ORDER BY id DESC')
    rows = cursor.fetchall()
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prediction History"
    
    headers = ["ID", "Teks Dokumen", "Prediksi Topik", "Confidence (%)", "Waktu Waktu (s)", "Tanggal"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    for r in rows:
        ws.append([
            r['id'],
            r['text'],
            r['predicted_label'],
            round(r['confidence'] * 100, 2),
            round(r['execution_time'], 4),
            r['created_at']
        ])
        
    for row in ws.iter_rows(min_row=2, max_col=6):
        ws.cell(row=row[0].row, column=1).alignment = align_center
        ws.cell(row=row[0].row, column=3).alignment = align_center
        ws.cell(row=row[0].row, column=4).alignment = align_center
        ws.cell(row=row[0].row, column=5).alignment = align_center
        ws.cell(row=row[0].row, column=6).alignment = align_center
        
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 22
    
    wb.save(output_path)
    return output_path
